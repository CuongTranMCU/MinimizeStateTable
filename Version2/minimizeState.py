import time
start_time = time.time()

from flask import Flask, render_template
app = Flask(__name__)

import openpyxl
from openpyxl import Workbook, load_workbook


def findNextGroup():
    for i in range(len(data)):
        form = "{} {}"
        temp1 = data[i][0]
        temp2 = data[i][1]
        temp3 = data[i][2]
        nextGroup[temp1] = form.format(groupList[temp2],groupList[temp3])
    
    
def divideGroup(groupNum, stateCounter):
    checkSet = set()
    for temp in groupList:
        if (groupList[temp] == groupNum):
            checkSet.add(nextGroup[temp])

    # All the states in group have the same nextState 
    if (len(checkSet) == 1):
        return stateCounter
    
    indexSet = 0
    for temp in checkSet:
        # Pass the first index to the second 
        if (indexSet == 0):
            indexSet += 1
            continue
        
        stateCounter += 1
        for temp1 in nextGroup:
            if (nextGroup[temp1] == temp and groupList[temp1] == groupNum):
                groupList[temp1] = stateCounter


    return stateCounter
    

def offset(target):
    if (len(target) == 1):
        return ord(target) - ord(data[0][0])
    else:
        temp1 = int(target[1:])
        temp2 = int(data[0][0][1:])
    return temp1 - temp2


book = load_workbook('F:/Python/Project/Tuan_code/autoTestCase/autoTest25.xlsx')
#book = load_workbook('F:/Python/TestCase/bai12.xlsx')
sheet = book["Sheet1"]

rowAddress = ["A{}","B{}","C{}","D{}","E{}"]
index = 3
data = []

#Create empty set of groups
outputSet = set()

#Common format for cell in excel file
cellA = rowAddress[0].format(index)

#Load data from excel file
while (sheet[cellA].value != None):
    temp = sheet[cellA].value
    temp1 = sheet[rowAddress[1].format(index)].value
    temp2 = sheet[rowAddress[2].format(index)].value
    temp3 = sheet[rowAddress[3].format(index)].value
    temp4 = sheet[rowAddress[4].format(index)].value

    output = "{} {}"
    temp5 = output.format(temp3,temp4)

    data.append((temp, temp1, temp2, temp5))                         #List of tuples 

    outputSet.add(temp5)

    index += 1
    cellA = rowAddress[0].format(index)
    

groupList = {}
#Convert set into list (order of set is not preserved) 
outputList = list(outputSet)

#Phân nhóm theo output 
for i in range(0,len(data)):
    for j in range(0,len(outputList)):
        if (data[i][3] == outputList[j]):
            groupList[data[i][0]] = j+1
            break


stateCounter = len(outputList)
#print(stateCounter)
#print(groupList)

nextGroup = {}
findNextGroup()
#print(nextGroup)


#Main:
done = 0
while (done == 0):
   count = 0
   temp = stateCounter

   for i in range(1,temp+1):
       stateCounter = divideGroup(i,stateCounter)
       findNextGroup()
       if (temp == stateCounter):
           count +=1

   if (count == temp):
       done = 1

#Dictionary of number_states and states 
result = {}
check = [0] * stateCounter

for temp in groupList:
    index = groupList[temp]
    if (check[index-1] == 0):
        result[index] = temp
        check[index-1] = 1

    if (check == [1] * stateCounter):
        break
sttIndex = 0
# print("STT   Present State","Next State","\t  Output", sep="\t")
# print("             ","X=0   X=1","X=0   X=1", sep="\t\t")
# print("-"*60)

dataAfter = []

for temp in result:
    sttIndex += 1
    stringGroup = nextGroup[result[temp]]
    spacePosition = stringGroup.find(" ")
    index0 = int(stringGroup[:spacePosition])
    index1 = int(stringGroup[spacePosition:])

    output0 = offset(result[temp])
    output0 = data[output0][3]
    output0 = output0[0] + ' ' + output0[2]
    # print(sttIndex," " *10 ,result[temp],"\t\t",result[index0]," " *3,result[index1]," " * 14,output0)
    # form = "{}" + " " *10 + "{}" + "\t\t"  + "{}" + " " *6 + "{}" + " " *17 + "{}"
    # form = form.format(sttIndex,result[temp],result[index0],result[index1],output0)
    dataAfter.append((result[temp],result[index0],result[index1],output0))
    # print(form)

end_time = time.time()

execution_time = end_time - start_time

@app.route('/')
def index():
    return render_template('index.html',before = data, after = dataAfter, lenBefore = len(data), lenAfter = len(dataAfter), executionTime = execution_time)

if __name__ == '__main__':
    app.run(debug=True)





#Summary of variables
# 1/ data -> chứa raw data từ file excel 
# 2/ outputSet -> chứa các giá trị output khác nhau
# 3/ outputList -> chuyển từ set sang list vì set không thể thay đổi giá trị
# 4/ groupList -> dictionary chứa số thứ tự của nhóm trạng thái sau khi chia nhóm
# 5/ nextGroup -> dictionary chứa số thứ tự của trạng thái kế tiếp